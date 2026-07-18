import json
import logging
import re
import time
import unicodedata
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models.entites import SectionProjet
from repositories.section_projet_repository import SectionProjetRepository


EXCLUDED_SHEETS = {"BD", "BD_INSTAL", "BD_VOIRIE", "BD_DECHETS", "BD_LOC"}
logger = logging.getLogger(__name__)


@dataclass
class DpgfSheetInfo:
    nom: str
    header_row: Optional[int] = None
    columns: Dict[str, int] = field(default_factory=dict)
    recognized: bool = False
    excluded: bool = False
    warning: Optional[str] = None


@dataclass
class DpgfImportSummary:
    feuilles_analysees: int = 0
    feuilles_lots_reconnues: List[str] = field(default_factory=list)
    feuilles_bd_exclues: List[str] = field(default_factory=list)
    feuilles_ignorees: List[str] = field(default_factory=list)
    conteneurs: int = 0
    ouvrages_chiffrables: int = 0
    ouvrages_pour_memoire: int = 0
    lignes_informatives: int = 0
    lignes_ignorees: int = 0
    numeros_articles_normalises: int = 0
    cellules_fusionnees_traitees: int = 0
    erreurs: List[str] = field(default_factory=list)
    avertissements: List[str] = field(default_factory=list)
    sections_importees: int = 0
    premiere_feuille_duree_secondes: float = 0.0
    duree_totale_secondes: float = 0.0


class ImportDpgfService:
    REQUIRED_COLUMNS = {
        "numero_article": {"nart", "numeroart", "noart", "nart.", "narticle", "numeroarticle"},
        "libelle": {"libelles", "libelle", "designation", "descriptif"},
        "unite": {"u", "unite", "unit"},
        "quantite": {"qtes", "qte", "quantite", "quantites"},
        "prix_unitaire": {"pu", "p.u", "prixunitaire", "prixunit"},
        "total": {"total", "montant"},
    }

    def __init__(self, section_repository: SectionProjetRepository):
        self.section_repository = section_repository

    def analyser_fichier(self, filepath: str) -> List[DpgfSheetInfo]:
        logger.info("Analyse DPGF : ouverture du fichier %s", filepath)
        workbook = load_workbook(filepath, data_only=False)
        infos = []
        logger.info("Analyse DPGF : %s feuilles détectées", len(workbook.sheetnames))
        for sheet_name in workbook.sheetnames:
            logger.info("Analyse DPGF : feuille '%s'", sheet_name)
            worksheet = workbook[sheet_name]
            info = DpgfSheetInfo(nom=sheet_name)
            if self._is_excluded_sheet(sheet_name):
                info.excluded = True
                info.warning = "Feuille BD exclue."
                logger.info("Analyse DPGF : feuille '%s' exclue", sheet_name)
            else:
                header = self.detect_header(worksheet)
                if header:
                    info.header_row = header["row"]
                    info.columns = header["columns"]
                    info.recognized = True
                    logger.info("Analyse DPGF : feuille '%s' reconnue, en-tête ligne %s", sheet_name, info.header_row)
                else:
                    info.warning = "Feuille ignorée : en-tête DPGF non détecté dans les 15 premières lignes."
                    logger.warning("Analyse DPGF : feuille '%s' ignorée, en-tête introuvable", sheet_name)
            infos.append(info)
        return infos

    def importer_fichier(
        self,
        filepath: str,
        projet_id: int,
        header_overrides: Optional[Dict[str, int]] = None,
        timeout_seconds: Optional[float] = 120,
    ) -> DpgfImportSummary:
        logger.info("Import DPGF : début projet_id=%s fichier=%s", projet_id, filepath)
        start_time = time.monotonic()
        summary = DpgfImportSummary()
        workbook = load_workbook(filepath, data_only=False)
        source_file = Path(filepath).name
        sections: List[SectionProjet] = []
        temp_counter = 0

        summary.feuilles_analysees = len(workbook.sheetnames)
        header_overrides = header_overrides or {}

        for sheet_name in workbook.sheetnames:
            self._check_timeout(start_time, timeout_seconds)
            logger.info("Import DPGF : traitement feuille '%s'", sheet_name)
            worksheet = workbook[sheet_name]
            if self._is_excluded_sheet(sheet_name):
                summary.feuilles_bd_exclues.append(sheet_name)
                logger.info("Import DPGF : feuille '%s' exclue", sheet_name)
                continue

            detected = self.detect_header(worksheet, header_overrides.get(sheet_name))
            if not detected:
                summary.feuilles_ignorees.append(sheet_name)
                summary.avertissements.append(
                    f"Feuille '{sheet_name}' ignorée : en-tête DPGF non détecté."
                )
                logger.warning("Import DPGF : feuille '%s' ignorée, en-tête introuvable", sheet_name)
                continue

            summary.feuilles_lots_reconnues.append(sheet_name)
            header_row = detected["row"]
            columns = detected["columns"]
            logger.info("Import DPGF : feuille '%s' reconnue avec en-tête ligne %s", sheet_name, header_row)
            logger.info(
                "Import DPGF : dimensions brutes feuille '%s' max_row=%s max_column=%s",
                sheet_name,
                worksheet.max_row,
                worksheet.max_column,
            )
            relevant_columns = sorted(set(columns.values()))
            data_rows = self._detect_data_rows(worksheet, header_row, relevant_columns)
            logger.info(
                "Import DPGF : feuille '%s' lignes utiles détectées=%s",
                sheet_name,
                len(data_rows),
            )
            logger.info("Import DPGF : feuille '%s' construction index cellules fusionnées", sheet_name)
            merged_lookup = self._build_merged_lookup(worksheet, data_rows, relevant_columns)
            logger.info(
                "Import DPGF : feuille '%s' cellules fusionnées indexées=%s",
                sheet_name,
                len(merged_lookup),
            )
            if merged_lookup:
                summary.cellules_fusionnees_traitees += len(merged_lookup)

            temp_counter += 1
            sheet_temp_id = -temp_counter
            sections.append(self._make_section(
                temp_id=sheet_temp_id,
                projet_id=projet_id,
                parent_id=None,
                type_ligne="lot",
                numero_article=None,
                numero_article_original=None,
                libelle=sheet_name,
                unite=None,
                quantite=None,
                prix_unitaire=None,
                total=None,
                pour_memoire=False,
                ordre_affichage=len(sections),
                profondeur=1,
                fichier_source=source_file,
                feuille_source=sheet_name,
                ligne_excel_source=header_row,
                formule_total=None,
                donnees_source_json={"feuille_source": sheet_name, "header_row": header_row},
            ))

            last_by_depth: Dict[int, int] = {1: sheet_temp_id}
            last_numbered_depth = 1
            last_numbered_id = sheet_temp_id

            sheet_start = time.monotonic()
            for processed_index, row_number in enumerate(data_rows, start=1):
                if processed_index == 1 or processed_index % 25 == 0:
                    logger.info(
                        "Import DPGF : feuille '%s' progression lignes %s/%s (ligne Excel %s)",
                        sheet_name,
                        processed_index,
                        len(data_rows),
                        row_number,
                    )
                self._check_timeout(start_time, timeout_seconds)
                parsed = self._parse_row(worksheet, row_number, columns, merged_lookup)
                if parsed["type_ligne"] == "ignoree":
                    summary.lignes_ignorees += 1
                    continue

                numero_original = parsed["numero_article_original"]
                numero_normalise = self.normalize_article_number(numero_original)
                if numero_original and numero_normalise and numero_original != numero_normalise:
                    summary.numeros_articles_normalises += 1

                if numero_normalise:
                    depth = self.article_depth(numero_normalise)
                    parent_id = last_by_depth.get(depth - 1, sheet_temp_id)
                    for existing_depth in list(last_by_depth.keys()):
                        if existing_depth >= depth:
                            last_by_depth.pop(existing_depth, None)
                else:
                    depth = last_numbered_depth + 1
                    parent_id = last_numbered_id

                type_ligne = parsed["type_ligne"]
                temp_counter += 1
                temp_id = -temp_counter

                section = self._make_section(
                    temp_id=temp_id,
                    projet_id=projet_id,
                    parent_id=parent_id,
                    type_ligne=type_ligne,
                    numero_article=numero_normalise,
                    numero_article_original=numero_original,
                    libelle=parsed["libelle"],
                    unite=parsed["unite"],
                    quantite=parsed["quantite"],
                    prix_unitaire=parsed["prix_unitaire"],
                    total=parsed["total"],
                    pour_memoire=parsed["pour_memoire"],
                    ordre_affichage=len(sections),
                    profondeur=depth,
                    fichier_source=source_file,
                    feuille_source=sheet_name,
                    ligne_excel_source=row_number,
                    formule_total=parsed["formule_total"],
                    donnees_source_json=parsed["source"],
                )
                sections.append(section)

                if type_ligne == "conteneur":
                    summary.conteneurs += 1
                elif type_ligne == "ouvrage":
                    summary.ouvrages_chiffrables += 1
                elif type_ligne == "pour_memoire":
                    summary.ouvrages_pour_memoire += 1
                elif type_ligne == "information":
                    summary.lignes_informatives += 1

                last_by_depth[depth] = temp_id
                if numero_normalise:
                    last_numbered_depth = depth
                    last_numbered_id = temp_id
            logger.info(
                "Import DPGF : feuille '%s' terminée en %.3fs",
                sheet_name,
                time.monotonic() - sheet_start,
            )
            sheet_duration = time.monotonic() - sheet_start
            if summary.premiere_feuille_duree_secondes == 0.0:
                summary.premiere_feuille_duree_secondes = sheet_duration

        try:
            self._promote_parent_sections_to_containers(sections, summary)
            logger.info("Import DPGF : remplacement transactionnel de %s sections", len(sections))
            self.section_repository.replace_for_projet(projet_id, sections)
        except Exception as exc:
            summary.erreurs.append(f"Import annulé : {exc}")
            logger.exception("Import DPGF : rollback après erreur")
            raise

        summary.sections_importees = len(sections)
        summary.duree_totale_secondes = time.monotonic() - start_time
        logger.info(
            "Import DPGF : fin projet_id=%s sections=%s durée=%.3fs",
            projet_id,
            summary.sections_importees,
            summary.duree_totale_secondes,
        )
        return summary

    def detect_header(self, worksheet: Worksheet, forced_row: Optional[int] = None) -> Optional[Dict[str, Any]]:
        rows = [forced_row] if forced_row else range(1, min(worksheet.max_row, 15) + 1)
        for row_number in rows:
            if not row_number:
                continue
            columns = {}
            for col_number in self._non_empty_columns_for_row(worksheet, row_number):
                value = worksheet.cell(row_number, col_number).value
                key = self._header_key(value)
                for canonical, variants in self.REQUIRED_COLUMNS.items():
                    if key in variants:
                        columns[canonical] = col_number
            if set(columns.keys()) >= set(self.REQUIRED_COLUMNS.keys()):
                return {"row": row_number, "columns": columns}
        return None

    def normalize_article_number(self, value: Any) -> Optional[str]:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        text = re.sub(r"\s*[.;,\s]\s*", ".", text)
        text = re.sub(r"\.+", ".", text).strip(".")
        return text or None

    def article_depth(self, value: str) -> int:
        return len([part for part in value.split(".") if part])

    def _parse_row(self, worksheet: Worksheet, row_number: int, columns: Dict[str, int], merged_lookup: Dict[tuple, tuple]) -> Dict[str, Any]:
        values = {
            key: self._cell_value(worksheet, row_number, col, merged_lookup)
            for key, col in columns.items()
        }
        source_cells = self._source_cells_for_row(worksheet, row_number, columns, merged_lookup)

        numero = self._text(values.get("numero_article"))
        libelle = self._text(values.get("libelle"))
        unite = self._text(values.get("unite"))
        total_value = values.get("total")
        formule_total = total_value if isinstance(total_value, str) and total_value.startswith("=") else None

        if not numero and not libelle and not unite and not formule_total:
            return {"type_ligne": "ignoree"}

        pour_memoire = unite.upper() == "PM"
        if pour_memoire:
            type_ligne = "pour_memoire"
        elif self._is_sum_formula(formule_total):
            type_ligne = "conteneur"
        elif self._is_multiplication_formula(formule_total) or unite:
            type_ligne = "ouvrage"
        else:
            type_ligne = "information"

        return {
            "type_ligne": type_ligne,
            "numero_article_original": numero,
            "libelle": libelle,
            "unite": unite or None,
            "quantite": self._decimal(values.get("quantite")),
            "prix_unitaire": self._decimal(values.get("prix_unitaire")),
            "total": self._decimal(values.get("total")),
            "pour_memoire": pour_memoire,
            "formule_total": formule_total,
            "source": {
                "cellules": source_cells,
                "numero_article_original": numero,
                "libelle": libelle,
                "unite": unite,
                "quantite": self._json_value(values.get("quantite")),
                "prix_unitaire": self._json_value(values.get("prix_unitaire")),
                "total": self._json_value(values.get("total")),
                "formule_total": formule_total,
                "ligne_excel_source": row_number,
            },
        }

    def _make_section(self, temp_id: int, **kwargs) -> SectionProjet:
        return SectionProjet(
            id=temp_id,
            date_creation="",
            date_modification="",
            donnees_source_json=json.dumps(kwargs.pop("donnees_source_json"), ensure_ascii=False),
            **kwargs,
        )

    def _promote_parent_sections_to_containers(self, sections: List[SectionProjet], summary: DpgfImportSummary):
        parent_ids = {section.parent_id for section in sections if section.parent_id is not None}
        for section in sections:
            if section.id in parent_ids and section.type_ligne not in ("lot", "conteneur", "pour_memoire"):
                if section.type_ligne == "ouvrage":
                    summary.ouvrages_chiffrables = max(0, summary.ouvrages_chiffrables - 1)
                elif section.type_ligne == "information":
                    summary.lignes_informatives = max(0, summary.lignes_informatives - 1)
                section.type_ligne = "conteneur"
                section.pour_memoire = False
                summary.conteneurs += 1

    def _is_excluded_sheet(self, sheet_name: str) -> bool:
        normalized = self._normalize_text(sheet_name)
        return normalized in {self._normalize_text(s) for s in EXCLUDED_SHEETS} or normalized.startswith("bd")

    def _header_key(self, value: Any) -> str:
        return self._normalize_text(value).replace(".", "")

    def _normalize_text(self, value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip().lower()
        text = "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")
        return re.sub(r"[^a-z0-9]+", "", text)

    def _detect_data_rows(self, worksheet: Worksheet, header_row: int, relevant_columns: List[int]) -> List[int]:
        rows = set()
        relevant = set(relevant_columns)
        for (row, col), cell in worksheet._cells.items():
            if row <= header_row or col not in relevant:
                continue
            if cell.value is not None and str(cell.value).strip() != "":
                rows.add(row)

        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.max_row <= header_row:
                continue
            if not any(merged_range.min_col <= col <= merged_range.max_col for col in relevant):
                continue
            anchor_value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
            if anchor_value is None or str(anchor_value).strip() == "":
                continue
            if merged_range.max_row - merged_range.min_row > 500:
                logger.warning(
                    "Import DPGF : plage fusionnée très haute ignorée partiellement (%s), seule la ligne ancre et les lignes déjà utiles seront traitées",
                    str(merged_range),
                )
                rows.add(merged_range.min_row)
                continue
            for row in range(max(header_row + 1, merged_range.min_row), merged_range.max_row + 1):
                rows.add(row)
        return sorted(rows)

    def _non_empty_columns_for_row(self, worksheet: Worksheet, row_number: int) -> List[int]:
        return sorted(
            col for (row, col), cell in worksheet._cells.items()
            if row == row_number and cell.value is not None and str(cell.value).strip() != ""
        )

    def _build_merged_lookup(self, worksheet: Worksheet, data_rows: List[int], relevant_columns: List[int]) -> Dict[tuple, tuple]:
        lookup = {}
        row_set = set(data_rows)
        col_set = set(relevant_columns)
        for merged_range in worksheet.merged_cells.ranges:
            anchor = (merged_range.min_row, merged_range.min_col)
            rows = [row for row in row_set if merged_range.min_row <= row <= merged_range.max_row]
            cols = [col for col in col_set if merged_range.min_col <= col <= merged_range.max_col]
            for row in rows:
                for col in cols:
                    lookup[(row, col)] = anchor
        return lookup

    def _cell_value(self, worksheet: Worksheet, row: int, col: int, merged_lookup: Dict[tuple, tuple]) -> Any:
        value = worksheet.cell(row, col).value
        if value is None and (row, col) in merged_lookup:
            anchor_row, anchor_col = merged_lookup[(row, col)]
            return worksheet.cell(anchor_row, anchor_col).value
        return value

    def _source_cells_for_row(self, worksheet: Worksheet, row_number: int, columns: Dict[str, int], merged_lookup: Dict[tuple, tuple]) -> Dict[str, Any]:
        source_cells = {}
        columns_to_capture = set(columns.values())
        for (row, col), cell in worksheet._cells.items():
            if row == row_number and cell.value is not None:
                columns_to_capture.add(col)
        for row, col in merged_lookup:
            if row == row_number:
                columns_to_capture.add(col)
        for col_number in sorted(columns_to_capture):
            column_letter = get_column_letter(col_number)
            source_cells[column_letter] = self._json_value(
                self._cell_value(worksheet, row_number, col_number, merged_lookup)
            )
        return source_cells

    def _is_sum_formula(self, formula: Optional[str]) -> bool:
        if not formula:
            return False
        text = formula.upper().replace(" ", "")
        return text.startswith("=SUM(") or text.startswith("=SOMME(")

    def _is_multiplication_formula(self, formula: Optional[str]) -> bool:
        if not formula:
            return False
        text = formula.upper().replace(" ", "")
        return text.startswith("=") and "*" in text and not self._is_sum_formula(formula)

    def _decimal(self, value: Any) -> Optional[Decimal]:
        if value is None:
            return None
        if isinstance(value, str) and value.startswith("="):
            return None
        text = str(value).strip().replace(" ", "").replace(",", ".")
        if not text:
            return None
        try:
            return Decimal(text)
        except InvalidOperation:
            return None

    def _text(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _json_value(self, value: Any) -> Any:
        if isinstance(value, Decimal):
            return float(value)
        return value

    def _check_timeout(self, start_time: float, timeout_seconds: Optional[float]):
        if timeout_seconds is not None and time.monotonic() - start_time > timeout_seconds:
            logger.error("Import DPGF : timeout après %.1f secondes", timeout_seconds)
            raise TimeoutError(f"Import DPGF interrompu après {timeout_seconds:.0f} secondes.")

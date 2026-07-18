import os
import time
from pathlib import Path
from decimal import Decimal
from types import SimpleNamespace

import pytest
from openpyxl import Workbook
from PySide6.QtCore import QEventLoop, QItemSelectionModel, QThread, QTimer, Qt
from PySide6.QtGui import QColor
from PySide6.QtWidgets import QApplication, QDialog, QMessageBox, QPushButton, QHeaderView, QRadioButton

from database.db_manager import DatabaseManager
from models.entites import Bibliotheque, OuvrageBibliotheque, Projet, SectionProjet
from repositories.bibliotheque_repository import BibliothequeRepository
from repositories.ouvrage_bibliotheque_repository import OuvrageBibliothequeRepository
from repositories.projet_repository import ProjetRepository
from repositories.section_projet_repository import SectionProjetRepository
from repositories.correspondance_dpgf_repository import CorrespondanceDpgfRepository
from repositories.parametre_repository import ParametreRepository
from repositories.version_projet_repository import VersionProjetRepository
from services.import_dpgf_service import ImportDpgfService
from services.import_dpgf_service import DpgfSheetInfo
from services.correspondance_service import CorrespondanceService
from services.chiffrage_projet_service import ChiffrageProjetService
from services.parametre_service import ParametreService
from services.version_projet_service import SOURCE_ACTUEL, VersionProjetService
from ui.pages.projets_page import (
    ChiffrageLigneDialog,
    ChiffrageTableDialog,
    ComparaisonVersionsDialog,
    DpgfImportWorker,
    MappingPageDialog,
    MatchingWorker,
    ProposalSelectionDialog,
    ProjetsPage,
    QuickOuvrageCreateDialog,
)
from ui.theme import APP_STYLESHEET, COLORS

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")


@pytest.fixture(scope="session")
def qapp():
    app = QApplication.instance() or QApplication([])
    return app


@pytest.fixture
def temp_db_manager(tmp_path):
    db_path = tmp_path / "test_worker.db"
    migrations_dir = Path(__file__).parent.parent / "database" / "migrations"
    return DatabaseManager(db_path=db_path, migrations_dir=migrations_dir)


@pytest.fixture
def projet_id(temp_db_manager):
    repo = ProjetRepository(temp_db_manager)
    return repo.create(Projet(
        id=None,
        nom="Projet Worker",
        client="Client",
        reference="REF",
        statut="Nouveau",
        date_creation="",
        date_modification="",
    ))


def make_workbook(tmp_path, name="worker.xlsx", sheets=1):
    wb = Workbook()
    for index in range(sheets):
        ws = wb.active if index == 0 else wb.create_sheet(f"Lot {index + 1:02d}")
        ws.title = f"Lot {index + 1:02d}"
        headers = ["N° Art.", "Libellés", "U.", "Qtés", "P.U.", "Total"]
        for col, value in enumerate(headers, start=1):
            ws.cell(3, col).value = value
        ws["A4"] = str(index + 1)
        ws["B4"] = f"Chapitre {index + 1}"
        ws["F4"] = "=SUM(F5:F5)"
        ws["A5"] = f"{index + 1}.1"
        ws["B5"] = f"Ouvrage {index + 1}"
        ws["C5"] = "u"
        ws["D5"] = 1
        ws["E5"] = 2
        ws["F5"] = "=D5*E5"
    path = tmp_path / name
    wb.save(path)
    return path


def make_library_ouvrage(bibliotheque_id, code, designation):
    return OuvrageBibliotheque(
        id=None,
        bibliotheque_id=bibliotheque_id,
        code=code,
        designation=designation,
        famille="Cloisons",
        unite="u",
        mode_chiffrage="importe",
        fournitures_ht_import=Decimal("10"),
        mo_heures_import=Decimal("1"),
        taux_horaire_import=Decimal("40"),
        mo_ht_import=Decimal("40"),
        materiel_ht_import=Decimal("0"),
        transport_ht_import=Decimal("0"),
        sous_traitance_ht_import=Decimal("0"),
        debourse_sec_import=Decimal("50"),
        pv_st_ht_import=Decimal("60"),
        pv_eg_ht_import=Decimal("75"),
        debourse_sec_calcule=None,
        pv_st_ht_calcule=None,
        pv_eg_ht_calcule=None,
        source_calcul="importe",
        date_dernier_calcul=None,
        attributs_techniques="{}",
        donnees_source_json="{}",
        actif=True,
        date_creation="",
        date_modification="",
    )


def build_mapping_fixture(temp_db_manager, projet_id):
    section_repo = SectionProjetRepository(temp_db_manager)
    lot_id = section_repo.create(SectionProjet(
        None, projet_id, None, "lot", None, None, "Lot Cloisons", None,
        None, None, None, False, 0, 1, "dpgf.xlsx", "Lot Cloisons", 6, None, "{}", "", ""
    ))
    section_a = section_repo.create(SectionProjet(
        None, projet_id, lot_id, "ouvrage", "1.1", "1.1", "Cloison standard",
        "u", Decimal("1"), None, None, False, 1, 2, "dpgf.xlsx", "Lot Cloisons", 7, "=D7*E7", "{}", "", ""
    ))
    section_b = section_repo.create(SectionProjet(
        None, projet_id, lot_id, "ouvrage", "1.2", "1.2", "Doublage cloison",
        "u", Decimal("2"), None, None, False, 2, 2, "dpgf.xlsx", "Lot Cloisons", 8, "=D8*E8", "{}", "", ""
    ))
    biblio_id = BibliothequeRepository(temp_db_manager).create(Bibliotheque(None, "Cloisons", "", "Cloisons", True, "", ""))
    ouvrage_repo = OuvrageBibliothequeRepository(temp_db_manager)
    ob_a = ouvrage_repo.create(make_library_ouvrage(biblio_id, "CL-A", "Cloison standard"))
    ob_b = ouvrage_repo.create(make_library_ouvrage(biblio_id, "CL-B", "Doublage cloison"))
    corr_repo = CorrespondanceDpgfRepository(temp_db_manager)
    corr_a = corr_repo.upsert_proposition(section_a, ob_a, Decimal("91"))
    corr_b = corr_repo.upsert_proposition(section_b, ob_b, Decimal("88"))
    service = CorrespondanceService(
        temp_db_manager,
        corr_repo,
        section_repo,
        ParametreService(ParametreRepository(temp_db_manager)),
    )
    projet = ProjetRepository(temp_db_manager).get_by_id(projet_id)
    sections = section_repo.get_by_projet(projet_id)
    return projet, sections, service, {"section_a": section_a, "section_b": section_b, "corr_a": corr_a, "corr_b": corr_b}


def wait_for_worker(worker, timeout_ms=5000):
    result = {"success": None, "error": None, "finished": 0}
    loop = QEventLoop()
    worker.succes.connect(lambda summary: result.update(success=summary))
    worker.erreur.connect(lambda message: result.update(error=message))

    def on_finished():
        result["finished"] += 1
        loop.quit()

    worker.termine.connect(on_finished)
    QTimer.singleShot(timeout_ms, loop.quit)
    worker.run()
    loop.processEvents()
    return result


def test_worker_emet_success_et_finished(qapp, temp_db_manager, projet_id, tmp_path):
    repo = SectionProjetRepository(temp_db_manager)
    path = make_workbook(tmp_path)
    worker = DpgfImportWorker(temp_db_manager.db_path, temp_db_manager.migrations_dir, str(path), projet_id, timeout_seconds=10)

    result = wait_for_worker(worker)

    assert result["error"] is None
    assert result["success"] is not None
    assert result["finished"] == 1
    assert repo.count_by_projet(projet_id) == 3


def test_worker_emet_error_et_finished(qapp, temp_db_manager, projet_id, tmp_path):
    worker = DpgfImportWorker(temp_db_manager.db_path, temp_db_manager.migrations_dir, str(tmp_path / "missing.xlsx"), projet_id, timeout_seconds=10)

    result = wait_for_worker(worker)

    assert result["success"] is None
    assert result["error"]
    assert result["finished"] == 1


def test_confirmation_entete_declenche_un_seul_import(qapp, monkeypatch):
    calls = []
    dialog_calls = []
    page = ProjetsPage.__new__(ProjetsPage)

    sheets = [
        DpgfSheetInfo(nom="Lot 01", header_row=3, recognized=True),
        DpgfSheetInfo(nom="Notes", recognized=False, warning="en-tête introuvable"),
    ]

    def fake_start(filepath, projet, overrides):
        calls.append((filepath, projet, overrides))

    def fake_get_text(*args, **kwargs):
        dialog_calls.append((args, kwargs))
        return "Lot 01=3", True

    monkeypatch.setattr("ui.pages.projets_page.QInputDialog.getText", fake_get_text)
    monkeypatch.setattr(page, "_start_import_worker", fake_start)
    projet = type("Projet", (), {"id": 1})()
    overrides = ProjetsPage._resolve_header_overrides(page, sheets)
    page._start_import_worker("file.xlsx", projet, overrides)

    assert len(dialog_calls) == 1
    assert len(calls) == 1
    assert calls[0][2] == {"Lot 01": 3}


def test_absence_recursion_import_worker(qapp, monkeypatch):
    page = ProjetsPage.__new__(ProjetsPage)
    page.import_in_progress = True
    called = {"start": 0}

    def fake_start(*args):
        called["start"] += 1

    monkeypatch.setattr(page, "_start_import_worker", fake_start)
    ProjetsPage.on_import_dpgf(page)

    assert called["start"] == 0


def test_start_worker_refuse_double_start_sans_creer_nouveau_thread(qapp):
    page = ProjetsPage.__new__(ProjetsPage)
    existing_thread = object()
    existing_worker = object()
    page.import_in_progress = True
    page._thread = existing_thread
    page._worker = existing_worker

    ProjetsPage._start_import_worker(page, "file.xlsx", type("Projet", (), {"id": 1})(), {})

    assert page._thread is existing_thread
    assert page._worker is existing_worker


def test_bouton_reactive_apres_erreur(qapp):
    page = ProjetsPage.__new__(ProjetsPage)
    page.btn_import = type("Button", (), {
        "enabled": True,
        "setEnabled": lambda self, value: setattr(self, "enabled", value),
    })()
    page.import_in_progress = True
    page.cursor_overridden = False
    page.selected_project = lambda: object()

    page._set_import_running(False)

    assert page.import_in_progress is False
    assert page.btn_import.enabled is True


def test_import_plusieurs_feuilles_sans_boucle(qapp, temp_db_manager, projet_id, tmp_path):
    path = make_workbook(tmp_path, sheets=3)
    worker = DpgfImportWorker(temp_db_manager.db_path, temp_db_manager.migrations_dir, str(path), projet_id, timeout_seconds=10)

    result = wait_for_worker(worker)

    assert result["error"] is None
    assert result["finished"] == 1
    assert result["success"].feuilles_lots_reconnues == ["Lot 01", "Lot 02", "Lot 03"]


def test_import_ui_success_message_et_arbre_sur_thread_principal(qapp, temp_db_manager, projet_id, tmp_path, monkeypatch):
    page = ProjetsPage()
    page.db_manager = temp_db_manager
    page.projet_repo = ProjetRepository(temp_db_manager)
    page.projet_service.repository = page.projet_repo
    page.section_repo = SectionProjetRepository(temp_db_manager)
    page.section_service.repository = page.section_repo
    page.import_service = ImportDpgfService(page.section_repo)
    page.load_projects()
    page.projects_table.selectRow(0)

    messages = []
    monkeypatch.setattr(
        "ui.pages.projets_page.QMessageBox.information",
        lambda *args, **kwargs: messages.append((args, kwargs)),
    )

    path = make_workbook(tmp_path, sheets=2)
    projet = page.selected_project()
    loop = QEventLoop()
    page._start_import_worker(str(path), projet, {})
    page.import_thread.finished.connect(loop.quit)
    QTimer.singleShot(5000, loop.quit)
    loop.exec()

    assert page.import_in_progress is False
    assert messages
    assert page.tree.topLevelItemCount() == 2
    assert page.btn_import.isEnabled() is True


def test_consultation_dpgf_modes_recherche_filtres_tooltip(qapp, temp_db_manager, projet_id, tmp_path):
    repo = SectionProjetRepository(temp_db_manager)
    service = ImportDpgfService(repo)
    path = make_workbook(tmp_path, sheets=1)
    service.importer_fichier(str(path), projet_id, timeout_seconds=10)

    page = ProjetsPage()
    page.db_manager = temp_db_manager
    page.projet_repo = ProjetRepository(temp_db_manager)
    page.projet_service.repository = page.projet_repo
    page.section_repo = repo
    page.section_service.repository = repo
    page.correspondance_repo = CorrespondanceDpgfRepository(temp_db_manager)
    page.parametre_service = ParametreService(ParametreRepository(temp_db_manager))
    page.correspondance_service = CorrespondanceService(
        temp_db_manager,
        page.correspondance_repo,
        page.section_repo,
        page.parametre_service,
    )
    page.load_projects()
    page.projects_table.selectRow(0)
    page.open_project(page.selected_project())

    lot_item = page.tree.topLevelItem(0)
    page.tree.setCurrentItem(lot_item)
    page.view_mode_combo.setCurrentText("Enfants directs")
    assert page.children_table.columnCount() == 11
    assert page.children_table.horizontalHeaderItem(8).text() == "Ligne Excel"
    assert page.children_table.horizontalHeaderItem(9).text() == "Correspondance"
    assert page.children_table.horizontalHeaderItem(10).text() == "Action"
    assert page.children_table.rowCount() == 1

    page.view_mode_combo.setCurrentText("Tous les ouvrages descendants")
    assert page.children_table.rowCount() == 1
    assert page.children_table.item(0, 2).text() == "ouvrage"
    action_buttons = page.children_table.cellWidget(0, 10).findChildren(QPushButton)
    assert [button.text() for button in action_buttons] == ["..."]

    page.dpgf_search_input.setText("Ouvrage 1")
    assert page.children_table.rowCount() == 1
    page.dpgf_search_input.setText("introuvable")
    assert page.children_table.rowCount() == 0
    page.dpgf_search_input.setText("")

    page.type_filter.setCurrentText("conteneur")
    assert page.children_table.rowCount() == 0
    page.type_filter.setCurrentText("Tous")
    page.unite_filter.setCurrentText("u")
    assert page.children_table.rowCount() == 1
    assert page.children_table.item(0, 7).toolTip() == "Lot 01"


def test_page_dpgf_valider_et_annuler_validation(qapp, temp_db_manager, projet_id, monkeypatch):
    projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    page = ProjetsPage()
    page.db_manager = temp_db_manager
    page.projet_repo = ProjetRepository(temp_db_manager)
    page.projet_service = page.projet_service.__class__(page.projet_repo)
    page.section_repo = SectionProjetRepository(temp_db_manager)
    page.section_service = page.section_service.__class__(page.section_repo)
    page.correspondance_repo = CorrespondanceDpgfRepository(temp_db_manager)
    page.parametre_service = ParametreService(ParametreRepository(temp_db_manager))
    page.correspondance_service = service
    page.chiffrage_service = ChiffrageProjetService(temp_db_manager, page.section_repo, page.correspondance_repo)
    page.open_project(projet)
    page.tree.setCurrentItem(page.tree.topLevelItem(0))
    page.view_mode_combo.setCurrentText("Tous les ouvrages descendants")

    selected_ids = []

    def fake_exec(dialog):
        selected_ids.append(dialog.table.item(0, 1).text())
        dialog.selected_correspondance_id = ids["corr_a"]
        return QDialog.Accepted

    monkeypatch.setattr("ui.pages.projets_page.ProposalSelectionDialog.exec", fake_exec)
    page.valider_proposition_dpgf(ids["section_a"])

    assert selected_ids == ["CL-A"]
    assert service.statut_ouvrage(ids["section_a"]) == "Validée"
    buttons = page.children_table.cellWidget(0, 10).findChildren(QPushButton)
    assert [button.text() for button in buttons] == ["Annuler", "..."]

    page.annuler_validation_dpgf(ids["section_a"])

    assert service.statut_ouvrage(ids["section_a"]) == "Proposée"
    buttons = page.children_table.cellWidget(0, 10).findChildren(QPushButton)
    assert [button.text() for button in buttons] == ["Valider", "..."]


def test_fenetre_propositions_large_et_lisible(qapp, temp_db_manager, projet_id):
    _projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    propositions = service.correspondances_pour_ouvrage(ids["section_a"])

    dialog = ProposalSelectionDialog(propositions)

    assert dialog.minimumWidth() >= 900
    assert dialog.table.horizontalHeader().sectionResizeMode(2) == QHeaderView.Stretch
    assert dialog.table.item(0, 2).text()


def test_mapping_page_affiche_toutes_lignes_et_propositions(qapp, temp_db_manager, projet_id):
    projet, sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    dialog = MappingPageDialog(projet, sections, service)

    radios = dialog.findChildren(QRadioButton)
    assert len(radios) >= 2
    assert dialog.empty_label.isVisible() is False


def test_mapping_propositions_exclusives_par_ligne(qapp, temp_db_manager, projet_id):
    projet, sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    biblio_id = BibliothequeRepository(temp_db_manager).create(Bibliotheque(None, "Cloisons bis", "", "Cloisons", True, "", ""))
    extra = OuvrageBibliothequeRepository(temp_db_manager).create(
        make_library_ouvrage(biblio_id, "CL-X", "Cloison variante")
    )
    CorrespondanceDpgfRepository(temp_db_manager).upsert_proposition(ids["section_a"], extra, Decimal("70"))
    dialog = MappingPageDialog(projet, SectionProjetRepository(temp_db_manager).get_by_projet(projet_id), service)

    radios = [radio for radio in dialog.findChildren(QRadioButton) if f"_{ids['section_a']}_" in radio.objectName()]
    assert len(radios) >= 2
    radios[0].setChecked(True)
    radios[1].setChecked(True)

    assert sum(1 for radio in radios if radio.isChecked()) == 1


def test_mapping_valide_une_seule_proposition_meme_bibliotheque(qapp, temp_db_manager, projet_id):
    projet, sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    corr_repo = CorrespondanceDpgfRepository(temp_db_manager)
    biblio_id = BibliothequeRepository(temp_db_manager).get_all()[0].id
    ouvrage_2 = OuvrageBibliothequeRepository(temp_db_manager).create(
        make_library_ouvrage(biblio_id, "CL-ALT", "Cloison alternative même bibliothèque")
    )
    corr_2 = corr_repo.upsert_proposition(ids["section_a"], ouvrage_2, Decimal("79"))
    dialog = MappingPageDialog(projet, SectionProjetRepository(temp_db_manager).get_by_projet(projet_id), service)

    dialog.selected_by_section[ids["section_a"]] = corr_2
    dialog.validate_one(ids["section_a"])

    correspondances = service.correspondances_pour_ouvrage(ids["section_a"])
    validees = [c for c in correspondances if c["statut"] == "validee"]
    assert len(validees) == 1
    assert validees[0]["id"] == corr_2
    assert all(c["statut"] != "validee" for c in correspondances if c["id"] != corr_2)


def test_vue_fusionnee_validation_proposition_copie_chiffrage(qapp, temp_db_manager, projet_id, monkeypatch):
    projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    chiffrage = ChiffrageProjetService(
        temp_db_manager,
        SectionProjetRepository(temp_db_manager),
        CorrespondanceDpgfRepository(temp_db_manager),
    )
    dialog = ChiffrageTableDialog(
        projet,
        chiffrage,
        correspondance_service=service,
        db_manager=temp_db_manager,
    )
    ouvrage = next(row for row in dialog.rows if row["section_id"] == ids["section_a"])
    row = dialog.row_by_ouvrage_id[ouvrage["id"]]
    combo = dialog.table.cellWidget(row, dialog.mapping_col_proposals)
    monkeypatch.setattr("ui.pages.projets_page.QMessageBox.question", lambda *args, **kwargs: QMessageBox.Yes)
    combo.setCurrentIndex(combo.findData(ids["corr_a"]))

    updated = next(row for row in dialog.rows if row["id"] == ouvrage["id"])
    assert service.statut_ouvrage(ids["section_a"]) == "Validée"
    assert updated["ds_mat"] == Decimal("10")
    assert updated["ds_mo"] == Decimal("40")
    assert updated["ds_total"] == Decimal("50")
    assert updated["pv_total"] == Decimal("75")
    assert dialog.table.item(row, dialog.mapping_col_status).text() == "Validée"
    assert dialog.dashboard_labels["ds_total"].text() == "50.00 €"


def test_vue_fusionnee_choix_aucune_proposition_annule_validation(qapp, temp_db_manager, projet_id, monkeypatch):
    projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    chiffrage = ChiffrageProjetService(
        temp_db_manager,
        SectionProjetRepository(temp_db_manager),
        CorrespondanceDpgfRepository(temp_db_manager),
    )
    dialog = ChiffrageTableDialog(
        projet,
        chiffrage,
        correspondance_service=service,
        db_manager=temp_db_manager,
    )
    ouvrage = next(row for row in dialog.rows if row["section_id"] == ids["section_a"])
    row = dialog.row_by_ouvrage_id[ouvrage["id"]]
    combo = dialog.table.cellWidget(row, dialog.mapping_col_proposals)
    monkeypatch.setattr("ui.pages.projets_page.QMessageBox.question", lambda *args, **kwargs: QMessageBox.Yes)
    combo.setCurrentIndex(combo.findData(ids["corr_a"]))
    assert service.statut_ouvrage(ids["section_a"]) == "Validée"

    combo = dialog.table.cellWidget(row, dialog.mapping_col_proposals)
    combo.setCurrentIndex(0)
    updated = next(row_data for row_data in dialog.rows if row_data["id"] == ouvrage["id"])

    assert combo.currentText() == "Aucune proposition"
    assert service.statut_ouvrage(ids["section_a"]) == "Aucune"
    assert updated["ds_total"] == Decimal("0.00")
    assert updated["pv_total"] == Decimal("0.00")
    assert dialog.dashboard_labels["ds_total"].text() == "0.00 €"
    assert dialog.dashboard_labels["pv_total"].text() == "0.00 €"
    assert dialog.table.item(row, dialog.mapping_col_status).text() == "Aucune"
    assert dialog.table.item(row, dialog.mapping_col_status).background().color() == QColor(COLORS["danger"])


def test_vue_fusionnee_choix_aucune_proposition_peut_etre_refuse(qapp, temp_db_manager, projet_id, monkeypatch):
    projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    chiffrage = ChiffrageProjetService(
        temp_db_manager,
        SectionProjetRepository(temp_db_manager),
        CorrespondanceDpgfRepository(temp_db_manager),
    )
    dialog = ChiffrageTableDialog(
        projet,
        chiffrage,
        correspondance_service=service,
        db_manager=temp_db_manager,
    )
    ouvrage = next(row for row in dialog.rows if row["section_id"] == ids["section_a"])
    row = dialog.row_by_ouvrage_id[ouvrage["id"]]
    combo = dialog.table.cellWidget(row, dialog.mapping_col_proposals)
    monkeypatch.setattr("ui.pages.projets_page.QMessageBox.question", lambda *args, **kwargs: QMessageBox.Yes)
    combo.setCurrentIndex(combo.findData(ids["corr_a"]))
    combo = dialog.table.cellWidget(row, dialog.mapping_col_proposals)

    monkeypatch.setattr("ui.pages.projets_page.QMessageBox.question", lambda *args, **kwargs: QMessageBox.No)
    combo.setCurrentIndex(0)

    assert combo.currentData() == ids["corr_a"]
    assert service.statut_ouvrage(ids["section_a"]) == "Validée"
    assert dialog.dashboard_labels["ds_total"].text() == "50.00 €"


def test_tableau_bord_chiffrage_temps_reel(qapp, temp_db_manager, projet_id):
    projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    chiffrage = ChiffrageProjetService(
        temp_db_manager,
        SectionProjetRepository(temp_db_manager),
        CorrespondanceDpgfRepository(temp_db_manager),
    )
    service.associer_resultat_pour_ouvrage(ids["section_a"], ids["corr_a"])
    chiffrage.copier_depuis_bibliotheque(ids["section_a"])
    dialog = ChiffrageTableDialog(
        projet,
        chiffrage,
        correspondance_service=service,
        db_manager=temp_db_manager,
    )
    ouvrage = next(row for row in dialog.rows if row["section_id"] == ids["section_a"])
    row = dialog.row_by_ouvrage_id[ouvrage["id"]]

    assert dialog.dashboard_labels["ds_total"].text() == "50.00 €"
    assert dialog.dashboard_labels["pv_total"].text() == "75.00 €"
    assert dialog.dashboard_labels["progress"].text() == "1 / 2"
    assert dialog.dashboard_labels["validated"].text() == "1"
    assert dialog.dashboard_labels["proposed"].text() == "1"
    assert dialog.dashboard_labels["manual"].text() == "0"
    assert dialog.dashboard_labels["untreated"].text() == "0"

    dialog.table.item(row, 4).setText("60")

    assert dialog.dashboard_labels["ds_total"].text() == "70.00 €"
    assert dialog.dashboard_labels["pv_total"].text() == "105.00 €"
    assert dialog.dashboard_labels["margin"].text() == "35.00 € / 33.33 %"
    assert dialog.nature_table.item(0, 1).text() == "60.00 €"
    assert dialog.nature_table.item(0, 2).text() == "85.71 %"
    assert dialog.group_totals_tree.topLevelItem(0).text(1) == "70.00 €"
    assert dialog.group_totals_tree.topLevelItem(0).child(0).text(2) == "105.00 €"

    proposed_ouvrage = next(row_data for row_data in dialog.rows if row_data["section_id"] == ids["section_b"])
    proposed_row = dialog.row_by_ouvrage_id[proposed_ouvrage["id"]]
    dialog.table.item(proposed_row, 4).setText("100")

    assert service.statut_ouvrage(ids["section_b"]) == "Proposée"
    assert dialog.dashboard_labels["ds_total"].text() == "70.00 €"
    assert dialog.dashboard_labels["pv_total"].text() == "105.00 €"
    assert dialog.nature_table.item(0, 1).text() == "60.00 €"


def test_table_mapping_chiffrage_designation_redimensionnable(qapp, temp_db_manager, projet_id):
    projet, _sections, service, _ids = build_mapping_fixture(temp_db_manager, projet_id)
    chiffrage = ChiffrageProjetService(
        temp_db_manager,
        SectionProjetRepository(temp_db_manager),
        CorrespondanceDpgfRepository(temp_db_manager),
    )
    dialog = ChiffrageTableDialog(
        projet,
        chiffrage,
        correspondance_service=service,
        db_manager=temp_db_manager,
    )

    header = dialog.table.horizontalHeader()
    assert header.sectionResizeMode(1) == QHeaderView.Interactive
    assert dialog.table.columnWidth(1) >= 300


def test_table_mapping_chiffrage_liste_propositions_elargie(qapp, temp_db_manager, projet_id):
    projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    chiffrage = ChiffrageProjetService(
        temp_db_manager,
        SectionProjetRepository(temp_db_manager),
        CorrespondanceDpgfRepository(temp_db_manager),
    )
    dialog = ChiffrageTableDialog(
        projet,
        chiffrage,
        correspondance_service=service,
        db_manager=temp_db_manager,
    )
    ouvrage = next(row for row in dialog.rows if row["section_id"] == ids["section_a"])
    table_row = dialog.row_by_ouvrage_id[ouvrage["id"]]
    combo = dialog.table.cellWidget(table_row, dialog.mapping_col_proposals)

    assert combo.view().minimumWidth() >= 900


def test_page_mapping_chiffrage_sauvegarde_travail_en_version(qapp, temp_db_manager, projet_id, monkeypatch):
    projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    chiffrage = ChiffrageProjetService(
        temp_db_manager,
        SectionProjetRepository(temp_db_manager),
        CorrespondanceDpgfRepository(temp_db_manager),
    )
    version_service = VersionProjetService(VersionProjetRepository(temp_db_manager))
    service.associer_resultat_pour_ouvrage(ids["section_a"], ids["corr_a"])
    chiffrage.copier_depuis_bibliotheque(ids["section_a"])
    dialog = ChiffrageTableDialog(
        projet,
        chiffrage,
        correspondance_service=service,
        db_manager=temp_db_manager,
        version_service=version_service,
    )
    monkeypatch.setattr("ui.pages.projets_page.QInputDialog.getText", lambda *args, **kwargs: ("Version chantier 1", True))
    monkeypatch.setattr("ui.pages.projets_page.QMessageBox.information", lambda *args, **kwargs: None)

    dialog.save_current_work_as_version()

    versions = version_service.lister_versions(projet_id)
    assert versions[0].nom == "Version chantier 1"
    assert versions[0].projet_id == projet_id
    assert versions[0].nombre_lignes == 2
    assert dialog.version_combo.findData(versions[0].id) >= 0
    assert dialog.version_combo.currentData() is None
    comparison = version_service.comparer(projet_id, str(versions[0].id), SOURCE_ACTUEL)
    assert all(row["ecart_montant"] == 0 for row in comparison["lignes"])


def test_page_mapping_chiffrage_affiche_version_selectionnee_puis_original(qapp, temp_db_manager, projet_id):
    projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    chiffrage = ChiffrageProjetService(
        temp_db_manager,
        SectionProjetRepository(temp_db_manager),
        CorrespondanceDpgfRepository(temp_db_manager),
    )
    version_service = VersionProjetService(VersionProjetRepository(temp_db_manager))
    service.associer_resultat_pour_ouvrage(ids["section_a"], ids["corr_a"])
    chiffrage.copier_depuis_bibliotheque(ids["section_a"])
    dialog = ChiffrageTableDialog(
        projet,
        chiffrage,
        correspondance_service=service,
        db_manager=temp_db_manager,
        version_service=version_service,
    )
    version_id = version_service.creer_version(projet_id, "Version 1")
    dialog.refresh_version_controls()
    ouvrage = next(row for row in dialog.rows if row["section_id"] == ids["section_a"])
    row = dialog.row_by_ouvrage_id[ouvrage["id"]]
    dialog.table.item(row, 4).setText("60")
    assert dialog.table.item(row, 9).text() == "70.00"
    assert dialog.dashboard_labels["ds_total"].text() == "70.00 €"

    dialog.version_combo.setCurrentIndex(dialog.version_combo.findData(version_id))
    row = dialog.row_by_ouvrage_id[ouvrage["id"]]

    assert dialog.viewing_version_id == version_id
    assert dialog.table.item(row, 9).text() == "50.00"
    assert dialog.dashboard_labels["ds_total"].text() == "50.00 €"
    assert dialog.table.item(row, dialog.mapping_col_status).text() == "Validée"
    assert dialog.table.item(row, dialog.mapping_col_status).background().color() == QColor(COLORS["success"])
    assert dialog.table.item(row, 4).flags() & Qt.ItemIsEditable
    assert dialog.dashboard_labels["validated"].text() == "1"
    assert dialog.dashboard_labels["proposed"].text() == "1"

    dialog.table.item(row, 4).setText("80")

    assert dialog.table.item(row, 9).text() == "90.00"
    comparison = version_service.comparer(projet_id, str(version_id), SOURCE_ACTUEL)
    rows = {row["cle"]: row for row in comparison["lignes"]}
    assert rows["ds_total"]["reference"] == Decimal("90.00")
    assert rows["ds_total"]["comparee"] == Decimal("70.00")

    dialog.show_original_work()
    row = dialog.row_by_ouvrage_id[ouvrage["id"]]

    assert dialog.viewing_version_id is None
    assert dialog.version_combo.currentData() is None
    assert dialog.table.item(row, 9).text() == "70.00"
    assert dialog.dashboard_labels["ds_total"].text() == "70.00 €"
    assert dialog.table.item(row, 4).flags() & Qt.ItemIsEditable


def test_page_mapping_chiffrage_version_change_proposition_et_dashboard(qapp, temp_db_manager, projet_id, monkeypatch):
    projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    chiffrage = ChiffrageProjetService(
        temp_db_manager,
        SectionProjetRepository(temp_db_manager),
        CorrespondanceDpgfRepository(temp_db_manager),
    )
    version_service = VersionProjetService(VersionProjetRepository(temp_db_manager))
    service.associer_resultat_pour_ouvrage(ids["section_a"], ids["corr_a"])
    chiffrage.copier_depuis_bibliotheque(ids["section_a"])
    version_id = version_service.creer_version(projet_id, "Version 1")
    dialog = ChiffrageTableDialog(
        projet,
        chiffrage,
        correspondance_service=service,
        db_manager=temp_db_manager,
        version_service=version_service,
    )
    dialog.refresh_version_controls()
    dialog.version_combo.setCurrentIndex(dialog.version_combo.findData(version_id))
    ouvrage_b = next(row for row in dialog.rows if row["section_id"] == ids["section_b"])
    row_b = dialog.row_by_ouvrage_id[ouvrage_b["id"]]

    assert dialog.table.item(row_b, dialog.mapping_col_status).text() == "Proposée"
    assert dialog.table.item(row_b, dialog.mapping_col_status).background().color() == QColor(COLORS["warning"])
    assert dialog.dashboard_labels["ds_total"].text() == "50.00 €"
    assert dialog.dashboard_labels["validated"].text() == "1"
    assert dialog.dashboard_labels["proposed"].text() == "1"
    combo = dialog.table.cellWidget(row_b, dialog.mapping_col_proposals)
    assert combo.isEnabled() is True

    monkeypatch.setattr("ui.pages.projets_page.QMessageBox.question", lambda *args, **kwargs: QMessageBox.Yes)
    monkeypatch.setattr("ui.pages.projets_page.QMessageBox.critical", lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError(args[2])))
    combo.setCurrentIndex(combo.findData(ids["corr_b"]))

    assert service.statut_ouvrage(ids["section_b"]) == "Validée"
    assert dialog.table.item(row_b, dialog.mapping_col_status).text() == "Validée"
    assert dialog.table.item(row_b, dialog.mapping_col_status).background().color() == QColor(COLORS["success"])
    assert dialog.dashboard_labels["ds_total"].text() == "150.00 €"
    assert dialog.dashboard_labels["pv_total"].text() == "225.00 €"
    assert dialog.dashboard_labels["validated"].text() == "2"
    assert dialog.dashboard_labels["proposed"].text() == "0"

    current = version_service.comparer(projet_id, str(version_id), SOURCE_ACTUEL)
    rows = {row["cle"]: row for row in current["lignes"]}
    assert rows["ds_total"]["reference"] == Decimal("150.00")
    assert rows["ds_total"]["comparee"] == Decimal("50.00")


def test_page_mapping_chiffrage_auto_mapping_version_garde_validees(qapp, temp_db_manager, projet_id):
    projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    chiffrage = ChiffrageProjetService(
        temp_db_manager,
        SectionProjetRepository(temp_db_manager),
        CorrespondanceDpgfRepository(temp_db_manager),
    )
    version_service = VersionProjetService(VersionProjetRepository(temp_db_manager))
    service.associer_resultat_pour_ouvrage(ids["section_a"], ids["corr_a"])
    chiffrage.copier_depuis_bibliotheque(ids["section_a"])
    service.supprimer_correspondances_ouvrage(ids["section_b"])
    version_id = version_service.creer_version(projet_id, "Version 1")
    dialog = ChiffrageTableDialog(
        projet,
        chiffrage,
        correspondance_service=service,
        db_manager=temp_db_manager,
        version_service=version_service,
    )
    dialog.refresh_version_controls()
    dialog.version_combo.setCurrentIndex(dialog.version_combo.findData(version_id))
    row_a = dialog.row_by_ouvrage_id[next(row for row in dialog.rows if row["section_id"] == ids["section_a"])["id"]]
    row_b = dialog.row_by_ouvrage_id[next(row for row in dialog.rows if row["section_id"] == ids["section_b"])["id"]]

    assert dialog.btn_auto_search.isEnabled() is True
    assert dialog.table.item(row_a, dialog.mapping_col_status).text() == "Validée"
    assert dialog.table.item(row_b, dialog.mapping_col_status).text() == "Aucune"

    dialog.search_auto_for_all_missing()
    row_a = dialog.row_by_ouvrage_id[next(row for row in dialog.rows if row["section_id"] == ids["section_a"])["id"]]
    row_b = dialog.row_by_ouvrage_id[next(row for row in dialog.rows if row["section_id"] == ids["section_b"])["id"]]

    assert dialog.viewing_version_id == version_id
    assert service.statut_ouvrage(ids["section_a"]) == "Validée"
    assert service.statut_ouvrage(ids["section_b"]) == "Proposée"
    assert dialog.table.item(row_a, dialog.mapping_col_status).text() == "Validée"
    assert dialog.table.item(row_b, dialog.mapping_col_status).text() == "Proposée"
    assert dialog.table.item(row_b, dialog.mapping_col_status).background().color() == QColor(COLORS["warning"])
    assert dialog.dashboard_labels["validated"].text() == "1"
    assert dialog.dashboard_labels["proposed"].text() == "1"


def test_vue_fusionnee_creation_ouvrage_valide_et_recherche(qapp, temp_db_manager, projet_id):
    projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    chiffrage = ChiffrageProjetService(
        temp_db_manager,
        SectionProjetRepository(temp_db_manager),
        CorrespondanceDpgfRepository(temp_db_manager),
    )
    section = SectionProjetRepository(temp_db_manager).get_by_id(ids["section_a"])
    dialog = QuickOuvrageCreateDialog(temp_db_manager, section)
    dialog.designation_input.setText("Ouvrage manuel fusion unique")
    dialog.famille_input.setText("Famille fusion")
    dialog.unite_input.setText("u")
    dialog.ds_mat_input.setValue(11)
    dialog.ds_mo_input.setValue(22)

    dialog.accept()
    service.associer_manuellement(ids["section_a"], dialog.created_ouvrage_id)
    copied = chiffrage.copier_depuis_bibliotheque(ids["section_a"])
    found = service.recherche_catalogue_libre("manuel fusion unique")

    assert dialog.created_ouvrage_id is not None
    assert service.statut_ouvrage(ids["section_a"]) == "Validée"
    assert copied["ds_mat"] == Decimal("11.00")
    assert copied["ds_mo"] == Decimal("22.00")
    assert copied["ds_total"] == Decimal("33.00")
    assert any(result["ouvrage_bibliotheque_id"] == dialog.created_ouvrage_id for result in found)


def test_vue_fusionnee_compteur_progression(qapp, temp_db_manager, projet_id):
    projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    chiffrage = ChiffrageProjetService(
        temp_db_manager,
        SectionProjetRepository(temp_db_manager),
        CorrespondanceDpgfRepository(temp_db_manager),
    )
    dialog = ChiffrageTableDialog(
        projet,
        chiffrage,
        correspondance_service=service,
        db_manager=temp_db_manager,
    )
    assert dialog.progress_label.text() == "0/2 lignes reliées"

    service.associer_resultat_pour_ouvrage(ids["section_a"], ids["corr_a"])
    dialog.update_mapping_progress()

    assert dialog.progress_label.text() == "1/2 lignes reliées"


def test_proposition_version_originale_uniquement_a_progression_complete(qapp, temp_db_manager, projet_id, monkeypatch):
    projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    chiffrage = ChiffrageProjetService(
        temp_db_manager,
        SectionProjetRepository(temp_db_manager),
        CorrespondanceDpgfRepository(temp_db_manager),
    )
    calls = []

    def fake_prompt(self):
        calls.append(self.progress_label.text())

    monkeypatch.setattr(ChiffrageTableDialog, "prompt_original_version_creation", fake_prompt)
    dialog = ChiffrageTableDialog(
        projet,
        chiffrage,
        correspondance_service=service,
        db_manager=temp_db_manager,
    )
    assert calls == []

    service.associer_resultat_pour_ouvrage(ids["section_a"], ids["corr_a"])
    dialog.update_mapping_progress()
    assert calls == []

    service.associer_resultat_pour_ouvrage(ids["section_b"], ids["corr_b"])
    dialog.update_mapping_progress()
    dialog.update_mapping_progress()

    assert calls == ["2/2 lignes reliées"]


def test_creation_version_originale_snapshot_complet(qapp, temp_db_manager, projet_id):
    projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    section_repo = SectionProjetRepository(temp_db_manager)
    corr_repo = CorrespondanceDpgfRepository(temp_db_manager)
    chiffrage = ChiffrageProjetService(temp_db_manager, section_repo, corr_repo)
    version_service = VersionProjetService(VersionProjetRepository(temp_db_manager))
    for section_id, corr_id in ((ids["section_a"], ids["corr_a"]), (ids["section_b"], ids["corr_b"])):
        service.associer_resultat_pour_ouvrage(section_id, corr_id)
        chiffrage.copier_depuis_bibliotheque(section_id)
    dialog = ChiffrageTableDialog(
        projet,
        chiffrage,
        version_service=version_service,
    )
    dialog.original_version_prompt_shown = True

    version_id = dialog.create_original_version("Version originale")

    versions = version_service.lister_versions(projet_id)
    assert versions[0].id == version_id
    assert versions[0].nombre_lignes == 2
    comparison = version_service.comparer(projet_id, str(version_id), SOURCE_ACTUEL)
    rows = {row["cle"]: row for row in comparison["lignes"]}
    assert rows["ds_total"]["reference"] == Decimal("150")
    assert rows["pv_total"]["reference"] == Decimal("225")


def test_duplication_version_avertit_si_modifications_non_sauvegardees(qapp, monkeypatch):
    page = ProjetsPage.__new__(ProjetsPage)
    page.current_project = None
    page.selected_project = lambda: SimpleNamespace(id=42)
    page.refresh_versions_table = lambda: None
    recorded = {"warning": "", "duplicate": None}

    class FakeVersionService:
        def a_modifications_non_sauvegardees(self, projet_id):
            return projet_id == 42

        def lister_versions(self, projet_id):
            return [object()]

        def dupliquer_version(self, version_source_id, nom):
            recorded["duplicate"] = (version_source_id, nom)
            return 99

    class FakeMessageBox:
        AcceptRole = object()
        DestructiveRole = object()
        Cancel = object()

        def __init__(self, parent=None):
            self.buttons = []

        def setWindowTitle(self, title):
            pass

        def setText(self, text):
            recorded["warning"] = text

        def addButton(self, *args):
            button = object()
            self.buttons.append(button)
            return button

        def exec(self):
            pass

        def clickedButton(self):
            return self.buttons[1]

        @staticmethod
        def information(*args, **kwargs):
            pass

        @staticmethod
        def critical(*args, **kwargs):
            raise AssertionError("Aucune erreur ne doit être affichée.")

    page.version_service = FakeVersionService()
    monkeypatch.setattr("ui.pages.projets_page.QMessageBox", FakeMessageBox)
    monkeypatch.setattr("ui.pages.projets_page.QInputDialog.getText", lambda *args, **kwargs: ("Version 2", True))

    page.on_duplicate_version(12)

    assert recorded["warning"] == (
        "Les modifications actuelles non enregistrées seront perdues. "
        "Créer une version avec l'état actuel d'abord, ou continuer sans sauvegarder ?"
    )
    assert recorded["duplicate"] == (12, "Version 2")


def test_chiffrage_dpgf_modifie_actuel_pas_version_figee(qapp, temp_db_manager, projet_id):
    _projet, _sections, _service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    section_repo = SectionProjetRepository(temp_db_manager)
    corr_repo = CorrespondanceDpgfRepository(temp_db_manager)
    corr_repo.valider(ids["corr_a"])
    chiffrage = ChiffrageProjetService(temp_db_manager, section_repo, corr_repo)
    version_service = VersionProjetService(VersionProjetRepository(temp_db_manager))

    ouvrage = chiffrage.copier_depuis_bibliotheque(ids["section_a"])
    assert ouvrage["ds_total"] == Decimal("50.00")
    assert ouvrage["pv_total"] == Decimal("75.00")
    v1 = version_service.creer_version(projet_id, "Version initiale")

    chiffrage.sauvegarder_chiffrage(
        ids["section_a"],
        Decimal("10"),
        Decimal("20"),
        Decimal("3"),
        Decimal("2"),
        Decimal("5"),
        Decimal("1.50"),
    )
    v2 = version_service.creer_version(projet_id, "Version 2")

    comparison = version_service.comparer(projet_id, str(v1), str(v2))
    rows = {row["cle"]: row for row in comparison["lignes"]}
    assert rows["ds_total"]["reference"] == Decimal("50")
    assert rows["ds_total"]["comparee"] == Decimal("40")
    current = version_service.comparer(projet_id, str(v1), SOURCE_ACTUEL)
    current_rows = {row["cle"]: row for row in current["lignes"]}
    assert current_rows["pv_total"]["comparee"] == Decimal("60")


def test_ecrans_recents_utilisent_theme_contraste(qapp, temp_db_manager, projet_id):
    projet, sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    corr_repo = CorrespondanceDpgfRepository(temp_db_manager)
    corr_repo.valider(ids["corr_a"])
    section_repo = SectionProjetRepository(temp_db_manager)
    chiffrage_service = ChiffrageProjetService(temp_db_manager, section_repo, corr_repo)
    version_service = VersionProjetService(VersionProjetRepository(temp_db_manager))

    mapping_dialog = MappingPageDialog(projet, sections, service)
    chiffrage_dialog = ChiffrageLigneDialog(section_repo.get_by_id(ids["section_a"]), chiffrage_service)
    chiffrage_table = ChiffrageTableDialog(projet, chiffrage_service)
    comparaison_dialog = ComparaisonVersionsDialog(projet, version_service)

    assert COLORS["background"] != COLORS["text"]
    assert "background-color: #ffffff" not in APP_STYLESHEET.lower()
    for dialog in (mapping_dialog, chiffrage_dialog, chiffrage_table, comparaison_dialog):
        assert dialog.styleSheet() == APP_STYLESHEET


def test_page_comparaison_versions_actuel_et_liste_depliable(qapp, temp_db_manager, projet_id):
    projet, _sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    section_repo = SectionProjetRepository(temp_db_manager)
    corr_repo = CorrespondanceDpgfRepository(temp_db_manager)
    chiffrage = ChiffrageProjetService(temp_db_manager, section_repo, corr_repo)
    version_service = VersionProjetService(VersionProjetRepository(temp_db_manager))
    for section_id, corr_id in ((ids["section_a"], ids["corr_a"]), (ids["section_b"], ids["corr_b"])):
        service.associer_resultat_pour_ouvrage(section_id, corr_id)
        chiffrage.copier_depuis_bibliotheque(section_id)
    version_service.creer_version(projet_id, "Version de référence")
    ouvrage = next(row for row in chiffrage.lister_ouvrages_projet(projet_id) if row["section_id"] == ids["section_a"])
    chiffrage.sauvegarder_composants_ouvrage(
        ouvrage["id"],
        Decimal("60"),
        Decimal("10"),
        Decimal("0"),
        Decimal("0"),
        Decimal("0"),
    )

    dialog = ComparaisonVersionsDialog(projet, version_service)

    assert dialog.compare_combo.itemText(0) == "Version actuelle"
    assert dialog.impact_group.isCheckable() is True
    assert dialog.impact_table.isHidden() is True
    dialog.impact_group.setChecked(True)
    assert dialog.impact_table.isHidden() is False
    ecarts = [dialog.summary_table.item(row, 3).text() for row in range(dialog.summary_table.rowCount())]
    assert any(" / " in ecart and "€" in ecart for ecart in ecarts)


def test_table_chiffrage_cellule_recalcule_totaux_et_historise(qapp, temp_db_manager, projet_id):
    projet, _sections, _service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    section_repo = SectionProjetRepository(temp_db_manager)
    corr_repo = CorrespondanceDpgfRepository(temp_db_manager)
    corr_repo.valider(ids["corr_a"])
    chiffrage = ChiffrageProjetService(temp_db_manager, section_repo, corr_repo)
    chiffrage.copier_depuis_bibliotheque(ids["section_a"])
    dialog = ChiffrageTableDialog(projet, chiffrage)
    ouvrage = next(row for row in dialog.rows if row["section_id"] == ids["section_a"])
    row = dialog.row_by_ouvrage_id[ouvrage["id"]]

    dialog.table.item(row, 4).setText("12.50")

    updated = next(row_data for row_data in dialog.rows if row_data["id"] == ouvrage["id"])
    assert updated["ds_mo"] == Decimal("12.50")
    assert updated["ds_total"] == Decimal("22.50")
    assert updated["pv_total"] == Decimal("33.75")
    assert dialog.table.item(row, 9).text() == "22.50"
    assert dialog.table.item(row, 12).text() == "33.75"
    with temp_db_manager.get_connection() as conn:
        count = conn.execute(
            "SELECT COUNT(*) FROM historique_ouvrages_projet WHERE ouvrage_projet_id = ?",
            (ouvrage["id"],),
        ).fetchone()[0]
    assert count == 2


def test_table_chiffrage_rejette_saisie_negative_ou_non_numerique(qapp, temp_db_manager, projet_id):
    projet, _sections, _service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    section_repo = SectionProjetRepository(temp_db_manager)
    corr_repo = CorrespondanceDpgfRepository(temp_db_manager)
    chiffrage = ChiffrageProjetService(temp_db_manager, section_repo, corr_repo)
    ouvrage = chiffrage.obtenir_ou_creer_ouvrage(ids["section_a"])
    dialog = ChiffrageTableDialog(projet, chiffrage)
    row = dialog.row_by_ouvrage_id[ouvrage["id"]]

    dialog.table.item(row, 4).setText("-1")
    assert dialog.table.item(row, 4).text() == "0.00"
    dialog.table.item(row, 4).setText("abc")
    assert dialog.table.item(row, 4).text() == "0.00"


def test_table_chiffrage_copie_bibliotheque_selection_multiple(qapp, temp_db_manager, projet_id):
    projet, _sections, _service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    section_repo = SectionProjetRepository(temp_db_manager)
    corr_repo = CorrespondanceDpgfRepository(temp_db_manager)
    corr_repo.valider(ids["corr_a"])
    corr_repo.valider(ids["corr_b"])
    chiffrage = ChiffrageProjetService(temp_db_manager, section_repo, corr_repo)
    dialog = ChiffrageTableDialog(projet, chiffrage)
    rows = []
    for section_id in (ids["section_a"], ids["section_b"]):
        ouvrage = next(row for row in dialog.rows if row["section_id"] == section_id)
        rows.append(dialog.row_by_ouvrage_id[ouvrage["id"]])
    selection = dialog.table.selectionModel()
    for row in rows:
        index = dialog.table.model().index(row, 0)
        selection.select(index, QItemSelectionModel.Select | QItemSelectionModel.Rows)

    result = chiffrage.copier_depuis_bibliotheque_plusieurs(dialog._selected_section_ids())
    dialog.rows = chiffrage.lister_ouvrages_projet(projet_id)
    dialog.recalculate_totals()

    assert result == {"copiees": 2, "ignorees": 0}
    copied = [row for row in dialog.rows if row["section_id"] in (ids["section_a"], ids["section_b"])]
    assert sum(row["ds_total"] for row in copied) == Decimal("150")


def test_table_chiffrage_centaines_lignes_fluide(qapp, temp_db_manager, projet_id):
    section_repo = SectionProjetRepository(temp_db_manager)
    lot_id = section_repo.create(SectionProjet(
        None, projet_id, None, "lot", None, None, "Lot Volume", None,
        None, None, None, False, 0, 1, "dpgf.xlsx", "Lot Volume", 1, None, "{}", "", ""
    ))
    for index in range(300):
        section_repo.create(SectionProjet(
            None, projet_id, lot_id, "ouvrage", str(index), str(index), f"Ouvrage {index}",
            "u", Decimal("1"), None, None, False, index + 1, 2, "dpgf.xlsx", "Lot Volume", index + 2, None, "{}", "", ""
        ))
    projet = ProjetRepository(temp_db_manager).get_by_id(projet_id)
    chiffrage = ChiffrageProjetService(temp_db_manager, section_repo, CorrespondanceDpgfRepository(temp_db_manager))

    start = time.monotonic()
    dialog = ChiffrageTableDialog(projet, chiffrage)
    first_ouvrage = dialog.rows[0]
    dialog.table.item(dialog.row_by_ouvrage_id[first_ouvrage["id"]], 4).setText("1")
    elapsed = time.monotonic() - start

    assert len(dialog.rows) == 300
    assert elapsed < 3


def test_mapping_page_aucune_proposition_affiche_lancement(qapp, temp_db_manager, projet_id):
    section_repo = SectionProjetRepository(temp_db_manager)
    lot_id = section_repo.create(SectionProjet(
        None, projet_id, None, "lot", None, None, "Lot Cloisons", None,
        None, None, None, False, 0, 1, "dpgf.xlsx", "Lot Cloisons", 6, None, "{}", "", ""
    ))
    section_repo.create(SectionProjet(
        None, projet_id, lot_id, "ouvrage", "1.1", "1.1", "Sans proposition",
        "u", None, None, None, False, 1, 2, "dpgf.xlsx", "Lot Cloisons", 7, "=D7*E7", "{}", "", ""
    ))
    service = CorrespondanceService(
        temp_db_manager,
        CorrespondanceDpgfRepository(temp_db_manager),
        section_repo,
        ParametreService(ParametreRepository(temp_db_manager)),
    )
    projet = ProjetRepository(temp_db_manager).get_by_id(projet_id)
    dialog = MappingPageDialog(projet, section_repo.get_by_projet(projet_id), service)

    assert dialog.empty_label.isHidden() is False
    assert "Aucune proposition" in dialog.empty_label.text()
    assert dialog.btn_run_auto.isHidden() is False


def test_matching_worker_progression_et_fin(qapp, temp_db_manager, projet_id):
    projet, sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    # Supprime les propositions préexistantes pour forcer un traitement réel.
    with temp_db_manager.get_connection() as conn:
        conn.execute("DELETE FROM correspondances_dpgf")
        conn.commit()

    worker = MatchingWorker(temp_db_manager.db_path, temp_db_manager.migrations_dir, projet_id, False)
    thread = QThread()
    worker.moveToThread(thread)
    result = {"progress": [], "success": None, "finished": 0, "error": None}
    loop = QEventLoop()
    thread.started.connect(worker.run)
    worker.progression.connect(lambda current, total, message: result["progress"].append((current, total, message)))
    worker.succes.connect(lambda summary: result.update(success=summary))
    worker.erreur.connect(lambda message: result.update(error=message))
    worker.termine.connect(thread.quit)
    worker.termine.connect(worker.deleteLater)
    thread.finished.connect(lambda: result.update(finished=result["finished"] + 1))
    thread.finished.connect(thread.deleteLater)
    thread.finished.connect(loop.quit)

    thread.start()
    QTimer.singleShot(5000, loop.quit)
    loop.exec()

    assert result["error"] is None
    assert result["success"] is not None
    assert result["success"].total >= 2
    assert result["progress"]
    assert result["finished"] == 1


def test_mapping_validation_ligne_par_ligne(qapp, temp_db_manager, projet_id):
    projet, sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    dialog = MappingPageDialog(projet, sections, service)

    dialog.selected_by_section[ids["section_a"]] = ids["corr_a"]
    dialog.validate_one(ids["section_a"])

    assert service.statut_ouvrage(ids["section_a"]) == "Validée"


def test_mapping_validation_groupee_transaction(qapp, temp_db_manager, projet_id):
    projet, sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    dialog = MappingPageDialog(projet, sections, service)
    dialog.selected_by_section = {
        ids["section_a"]: ids["corr_a"],
        ids["section_b"]: ids["corr_b"],
    }
    dialog.validate_selected()

    assert service.statut_ouvrage(ids["section_a"]) == "Validée"
    assert service.statut_ouvrage(ids["section_b"]) == "Validée"


def test_mapping_validation_groupee_rejette_double_meme_ligne(qapp, temp_db_manager, projet_id):
    projet, sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    corr_repo = CorrespondanceDpgfRepository(temp_db_manager)
    biblio_id = BibliothequeRepository(temp_db_manager).get_all()[0].id
    other_ob = OuvrageBibliothequeRepository(temp_db_manager).create(make_library_ouvrage(biblio_id, "CL-C", "Autre cloison"))
    other_corr = corr_repo.upsert_proposition(ids["section_a"], other_ob, Decimal("75"))

    with pytest.raises(ValueError):
        service.associer_plusieurs([ids["corr_a"], other_corr])


def test_mapping_filtres_statut_lot_recherche(qapp, temp_db_manager, projet_id):
    projet, sections, service, ids = build_mapping_fixture(temp_db_manager, projet_id)
    dialog = MappingPageDialog(projet, sections, service)

    dialog.status_filter.setCurrentText("Proposée")
    assert len(dialog.filtered_sections()) == 2
    dialog.search_input.setText("Doublage")
    filtered = dialog.filtered_sections()
    assert len(filtered) == 1
    assert filtered[0].id == ids["section_b"]
    dialog.lot_filter.setCurrentText("Lot Cloisons")
    assert len(dialog.filtered_sections()) == 1

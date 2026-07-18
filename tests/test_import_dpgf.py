import json
import time
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from database.db_manager import DatabaseManager
from models.entites import Projet, SectionProjet
from repositories.projet_repository import ProjetRepository
from repositories.section_projet_repository import SectionProjetRepository
from services.import_dpgf_service import ImportDpgfService


@pytest.fixture
def temp_db_manager(tmp_path):
    db_path = tmp_path / "test_dpgf.db"
    migrations_dir = Path(__file__).parent.parent / "database" / "migrations"
    return DatabaseManager(db_path=db_path, migrations_dir=migrations_dir)


@pytest.fixture
def projet_id(temp_db_manager):
    repo = ProjetRepository(temp_db_manager)
    return repo.create(Projet(
        id=None,
        nom="Projet DPGF",
        client="Client",
        reference="REF",
        statut="Nouveau",
        date_creation="",
        date_modification="",
    ))


@pytest.fixture
def section_repo(temp_db_manager):
    return SectionProjetRepository(temp_db_manager)


@pytest.fixture
def import_service(section_repo):
    return ImportDpgfService(section_repo)


def add_header(ws, row=3):
    headers = ["N° Art.", "Libellés", "U.", "Qtés", "P.U.", "Total"]
    for col, value in enumerate(headers, start=1):
        ws.cell(row, col).value = value


def save_workbook(wb, tmp_path, name="dpgf.xlsx"):
    path = tmp_path / name
    wb.save(path)
    return path


def test_exclusion_feuilles_bd_et_rejet_sans_entete(import_service, tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "BD"
    ws["A1"] = "base"
    ws2 = wb.create_sheet("BD_INSTAL")
    ws2["A1"] = "base"
    ws3 = wb.create_sheet("Feuille sans entete")
    ws3["A1"] = "pas un dpgf"
    path = save_workbook(wb, tmp_path)

    infos = import_service.analyser_fichier(str(path))

    assert [info.nom for info in infos if info.excluded] == ["BD", "BD_INSTAL"]
    ignored = next(info for info in infos if info.nom == "Feuille sans entete")
    assert ignored.recognized is False
    assert "en-tête DPGF non détecté" in ignored.warning


def test_detection_feuille_lot(import_service, tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lot 01"
    add_header(ws, row=5)
    path = save_workbook(wb, tmp_path)

    infos = import_service.analyser_fichier(str(path))

    assert infos[0].recognized is True
    assert infos[0].header_row == 5
    assert infos[0].columns["numero_article"] == 1


def test_import_dpgf_hierarchie_formules_pm_sources(import_service, section_repo, projet_id, tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lot Menuiseries"
    add_header(ws, row=3)

    ws.merge_cells("B4:C4")
    ws["A4"] = "2"
    ws["B4"] = "MENUISERIES"
    ws["F4"] = "=SUM(F5:F12)"

    ws["A5"] = "2;15"
    ws["C5"] = "m2"
    ws["D5"] = 4
    ws["E5"] = 10
    ws["F5"] = "=D5*E5"

    ws["A6"] = "2.4.1"
    ws["B6"] = "Sous chapitre profond"
    ws["F6"] = "=SUM(F7:F8)"

    ws["A7"] = "2.4.1.3"
    ws["B7"] = "Ouvrage profond"
    ws["C7"] = "u"
    ws["D7"] = None
    ws["E7"] = 12
    ws["F7"] = "=D7*E7"

    ws["B8"] = "Escalier principal"
    ws["C8"] = "PM"

    ws["B9"] = "RDC"

    ws["A10"] = "3"
    ws["B10"] = "AUTRE CHAPITRE"
    ws["F10"] = "=SUM(F11:F12)"

    ws["A11"] = "3.1"
    ws["B11"] = "Prix unitaire vide"
    ws["C11"] = "ml"
    ws["D11"] = 8
    ws["E11"] = None
    ws["F11"] = "=D11*E11"

    ws["A12"] = "3.1.1.1"
    ws["B12"] = "Quantité vide"
    ws["C12"] = "m3"
    ws["D12"] = None
    ws["E12"] = 20
    ws["F12"] = "=D12*E12"

    path = save_workbook(wb, tmp_path)

    summary = import_service.importer_fichier(str(path), projet_id)
    sections = section_repo.get_by_projet(projet_id)

    assert summary.feuilles_analysees == 1
    assert summary.feuilles_lots_reconnues == ["Lot Menuiseries"]
    assert summary.conteneurs == 5
    assert summary.ouvrages_chiffrables == 2
    assert summary.ouvrages_pour_memoire == 1
    assert summary.lignes_informatives == 1
    assert summary.numeros_articles_normalises == 1
    assert summary.cellules_fusionnees_traitees > 0

    by_label = {section.libelle: section for section in sections}
    assert by_label["MENUISERIES"].type_ligne == "conteneur"
    assert by_label["MENUISERIES"].formule_total == "=SUM(F5:F12)"
    assert by_label["MENUISERIES"].ligne_excel_source == 4

    normalized = next(section for section in sections if section.numero_article == "2.15")
    assert normalized.numero_article_original == "2;15"
    assert normalized.type_ligne == "conteneur"
    assert normalized.quantite == Decimal("4")
    assert normalized.prix_unitaire == Decimal("10")
    assert normalized.formule_total == "=D5*E5"

    deep = by_label["Ouvrage profond"]
    assert deep.numero_article == "2.4.1.3"
    assert deep.profondeur == 4
    assert deep.quantite is None

    pm = by_label["Escalier principal"]
    assert pm.type_ligne == "pour_memoire"
    assert pm.pour_memoire is True
    assert pm.parent_id == by_label["Ouvrage profond"].id

    info = by_label["RDC"]
    assert info.type_ligne == "information"
    assert info.parent_id == by_label["Ouvrage profond"].id

    top_level = [section for section in sections if section.parent_id == by_label["Lot Menuiseries"].id]
    assert {section.numero_article for section in top_level} >= {"2", "3"}

    missing_pu = by_label["Prix unitaire vide"]
    assert missing_pu.prix_unitaire is None
    missing_qty = by_label["Quantité vide"]
    assert missing_qty.quantite is None

    source = json.loads(normalized.donnees_source_json)
    assert source["cellules"]["A"] == "2;15"
    assert source["formule_total"] == "=D5*E5"
    assert normalized.fichier_source == "dpgf.xlsx"
    assert normalized.feuille_source == "Lot Menuiseries"
    assert normalized.ligne_excel_source == 5


def test_ligne_avec_enfants_devient_conteneur(import_service, section_repo, projet_id, tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lot 01"
    add_header(ws)
    ws["A4"] = "1"
    ws["B4"] = "Chapitre sans SUM"
    ws["C4"] = "u"
    ws["F4"] = "=D4*E4"
    ws["A5"] = "1.1"
    ws["B5"] = "Ouvrage enfant"
    ws["C5"] = "u"
    ws["F5"] = "=D5*E5"
    path = save_workbook(wb, tmp_path, "children_container.xlsx")

    import_service.importer_fichier(str(path), projet_id)
    sections = section_repo.get_by_projet(projet_id)
    parent = next(section for section in sections if section.libelle == "Chapitre sans SUM")
    child = next(section for section in sections if section.libelle == "Ouvrage enfant")

    assert parent.type_ligne == "conteneur"
    assert child.type_ligne == "ouvrage"


def test_remplacement_dpgf_existant(import_service, section_repo, projet_id, tmp_path):
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "Lot 01"
    add_header(ws1)
    ws1["A4"] = "1"
    ws1["B4"] = "Ancien"
    ws1["F4"] = "=SUM(F5:F5)"
    path1 = save_workbook(wb1, tmp_path, "old.xlsx")
    import_service.importer_fichier(str(path1), projet_id)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Lot 02"
    add_header(ws2)
    ws2["A4"] = "2"
    ws2["B4"] = "Nouveau"
    ws2["F4"] = "=SUM(F5:F5)"
    path2 = save_workbook(wb2, tmp_path, "new.xlsx")
    import_service.importer_fichier(str(path2), projet_id)

    labels = [section.libelle for section in section_repo.get_by_projet(projet_id)]
    assert "Ancien" not in labels
    assert "Nouveau" in labels


def test_rollback_en_cas_erreur(section_repo, import_service, projet_id, tmp_path, monkeypatch):
    existing = SectionProjet(
        id=None,
        projet_id=projet_id,
        parent_id=None,
        type_ligne="lot",
        numero_article=None,
        numero_article_original=None,
        libelle="Structure existante",
        unite=None,
        quantite=None,
        prix_unitaire=None,
        total=None,
        pour_memoire=False,
        ordre_affichage=0,
        profondeur=1,
        fichier_source="old.xlsx",
        feuille_source="Ancien",
        ligne_excel_source=1,
        formule_total=None,
        donnees_source_json="{}",
        date_creation="",
        date_modification="",
    )
    section_repo.create(existing)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lot 01"
    add_header(ws)
    ws["A4"] = "1"
    ws["B4"] = "Nouveau"
    ws["F4"] = "=SUM(F5:F5)"
    path = save_workbook(wb, tmp_path, "rollback.xlsx")

    original_create = section_repo.create

    def failing_create(section, conn=None):
        if section.libelle == "Nouveau":
            raise RuntimeError("erreur simulée")
        return original_create(section, conn)

    monkeypatch.setattr(section_repo, "create", failing_create)

    with pytest.raises(RuntimeError):
        import_service.importer_fichier(str(path), projet_id)

    labels = [section.libelle for section in section_repo.get_by_projet(projet_id)]
    assert labels == ["Structure existante"]


def test_detection_numero_et_normalisation(import_service):
    assert import_service.normalize_article_number("2.4.1") == "2.4.1"
    assert import_service.article_depth("2.4.1") == 3
    assert import_service.normalize_article_number("2;15") == "2.15"


def test_import_ignore_dimensions_excel_gonflees(import_service, section_repo, projet_id, tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lot Etanchéité_Couverture"
    add_header(ws, row=6)
    ws["A7"] = "1"
    ws["B7"] = "Chapitre"
    ws["F7"] = "=SUM(F8:F8)"
    ws["A8"] = "1.1"
    ws["B8"] = "Ouvrage"
    ws["C8"] = "m2"
    ws["F8"] = "=D8*E8"
    ws.cell(10000, 500).value = "cellule hors DPGF"
    path = save_workbook(wb, tmp_path, "wide_dimension.xlsx")

    start = time.perf_counter()
    summary = import_service.importer_fichier(str(path), projet_id, timeout_seconds=None)
    duration = time.perf_counter() - start

    assert duration < 1.0
    assert summary.sections_importees == 3
    assert summary.lignes_ignorees == 0
    assert summary.premiere_feuille_duree_secondes > 0
    assert summary.duree_totale_secondes > 0
